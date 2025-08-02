import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import threading
import pandas as pd
import random

class ExcelManager:
    def __init__(self, file_path='user_data.xlsx'):
        self.file_path = file_path
        self.lock = threading.Lock()
        self.initialize_excel()
    
    def initialize_excel(self):
        """初始化Excel文件，如果不存在則創建"""
        try:
            if not os.path.exists(self.file_path):
                # 創建新的工作簿
                workbook = Workbook()
                
                # 創建 User Activity 工作表
                worksheet = workbook.active
                worksheet.title = "User_Activity"
                
                # 設置 User Activity 標題行
                headers = ['User ID', 'Created Time', 'Activity Type', 'Page', 'Details', 'Timestamp']
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col, value=header)
                
                # 調整列寬
                worksheet.column_dimensions['A'].width = 15  # User ID
                worksheet.column_dimensions['B'].width = 20  # Created Time
                worksheet.column_dimensions['C'].width = 15  # Activity Type
                worksheet.column_dimensions['D'].width = 15  # Page
                worksheet.column_dimensions['E'].width = 50  # Details
                worksheet.column_dimensions['F'].width = 20  # Timestamp
                
                # 創建 Sender Sheet
                sender_sheet = workbook.create_sheet("Sender_Actions")
                sender_headers = ['User ID', 'Created Time', 'Post ID', 'Original Comment', 'Emotion Analysis', 
                                'Recommended Memes', 'Chosen Meme', 'Claude Response', 'Timestamp']
                for col, header in enumerate(sender_headers, 1):
                    sender_sheet.cell(row=1, column=col, value=header)
                
                # 調整 Sender 列寬
                sender_sheet.column_dimensions['A'].width = 15  # User ID
                sender_sheet.column_dimensions['B'].width = 20  # Created Time
                sender_sheet.column_dimensions['C'].width = 15  # Post ID
                sender_sheet.column_dimensions['D'].width = 40  # Original Comment
                sender_sheet.column_dimensions['E'].width = 30  # Emotion Analysis
                sender_sheet.column_dimensions['F'].width = 50  # Recommended Memes
                sender_sheet.column_dimensions['G'].width = 30  # Chosen Meme
                sender_sheet.column_dimensions['H'].width = 50  # Claude Response
                sender_sheet.column_dimensions['I'].width = 20  # Timestamp
                
                # 創建 Receiver Sheet
                receiver_sheet = workbook.create_sheet("Receiver_Actions")
                receiver_headers = ['User ID', 'Created Time', 'Post ID', 'Recommended Meme', 'Clicked', 
                                  'Original Comment Viewed', 'Timestamp']
                for col, header in enumerate(receiver_headers, 1):
                    receiver_sheet.cell(row=1, column=col, value=header)
                
                # 調整 Receiver 列寬
                receiver_sheet.column_dimensions['A'].width = 15  # User ID
                receiver_sheet.column_dimensions['B'].width = 20  # Created Time
                receiver_sheet.column_dimensions['C'].width = 15  # Post ID
                receiver_sheet.column_dimensions['D'].width = 30  # Recommended Meme
                receiver_sheet.column_dimensions['E'].width = 10  # Clicked
                receiver_sheet.column_dimensions['F'].width = 40  # Original Comment Viewed
                receiver_sheet.column_dimensions['G'].width = 20  # Timestamp
                
                workbook.save(self.file_path)
                print(f"Excel文件已創建: {self.file_path}")
            else:
                print(f"Excel文件已存在: {self.file_path}")
        except Exception as e:
            print(f"初始化Excel文件時出錯: {e}")
    
    def add_user_entry(self, user_id, activity_type='login', page='index', details=''):
        """添加用戶進入記錄到 User Activity 表"""
        with self.lock:
            try:
                workbook = openpyxl.load_workbook(self.file_path)
                
                # 檢查是否存在 User_Activity 工作表，如果不存在則創建
                if 'User_Activity' not in workbook.sheetnames:
                    print("User_Activity 工作表不存在，正在創建...")
                    worksheet = workbook.create_sheet("User_Activity")
                    
                    # 設置標題行
                    headers = ['User ID', 'Created Time', 'Activity Type', 'Page', 'Details', 'Timestamp']
                    for col, header in enumerate(headers, 1):
                        worksheet.cell(row=1, column=col, value=header)
                    
                    # 調整列寬
                    worksheet.column_dimensions['A'].width = 15  # User ID
                    worksheet.column_dimensions['B'].width = 20  # Created Time
                    worksheet.column_dimensions['C'].width = 15  # Activity Type
                    worksheet.column_dimensions['D'].width = 15  # Page
                    worksheet.column_dimensions['E'].width = 50  # Details
                    worksheet.column_dimensions['F'].width = 20  # Timestamp
                else:
                    worksheet = workbook['User_Activity']
                
                # 找到下一個空行
                next_row = worksheet.max_row + 1
                
                # 獲取當前時間
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                timestamp = datetime.now().timestamp()
                
                # 寫入數據
                worksheet.cell(row=next_row, column=1, value=user_id)
                worksheet.cell(row=next_row, column=2, value=current_time)
                worksheet.cell(row=next_row, column=3, value=activity_type)
                worksheet.cell(row=next_row, column=4, value=page)
                worksheet.cell(row=next_row, column=5, value=details)
                worksheet.cell(row=next_row, column=6, value=timestamp)
                
                workbook.save(self.file_path)
                print(f"用戶活動已記錄: {user_id} - {activity_type} - {current_time}")
                return True
            except Exception as e:
                print(f"寫入Excel時出錯: {e}")
                return False
    
    def add_sender_action(self, user_id, post_id, original_comment, emotion_analysis, 
                         recommended_memes, chosen_meme='', claude_response=''):
        """添加發送者動作記錄"""
        with self.lock:
            try:
                workbook = openpyxl.load_workbook(self.file_path)
                worksheet = workbook['Sender_Actions']
                
                # 找到下一個空行
                next_row = worksheet.max_row + 1
                
                # 獲取當前時間
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                timestamp = datetime.now().timestamp()
                
                # 格式化推薦memes
                if isinstance(recommended_memes, list):
                    memes_str = ', '.join(recommended_memes)
                else:
                    memes_str = str(recommended_memes)
                
                # 寫入數據
                worksheet.cell(row=next_row, column=1, value=user_id)
                worksheet.cell(row=next_row, column=2, value=current_time)
                worksheet.cell(row=next_row, column=3, value=post_id)
                worksheet.cell(row=next_row, column=4, value=original_comment)
                worksheet.cell(row=next_row, column=5, value=emotion_analysis)
                worksheet.cell(row=next_row, column=6, value=memes_str)
                worksheet.cell(row=next_row, column=7, value=chosen_meme)
                worksheet.cell(row=next_row, column=8, value=claude_response)
                worksheet.cell(row=next_row, column=9, value=timestamp)
                
                workbook.save(self.file_path)
                print(f"發送者動作已記錄: {user_id} - {post_id} - {current_time}")
                return True
            except Exception as e:
                print(f"寫入發送者動作時出錯: {e}")
                return False
    
    def add_receiver_action(self, user_id, post_id, recommended_meme, clicked=False, 
                          original_comment_viewed=''):
        """添加觀察者動作記錄"""
        with self.lock:
            try:
                workbook = openpyxl.load_workbook(self.file_path)
                worksheet = workbook['Receiver_Actions']
                
                # 找到下一個空行
                next_row = worksheet.max_row + 1
                
                # 獲取當前時間
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                timestamp = datetime.now().timestamp()
                
                # 寫入數據
                worksheet.cell(row=next_row, column=1, value=user_id)
                worksheet.cell(row=next_row, column=2, value=current_time)
                worksheet.cell(row=next_row, column=3, value=post_id)
                worksheet.cell(row=next_row, column=4, value=recommended_meme)
                worksheet.cell(row=next_row, column=5, value=clicked)
                worksheet.cell(row=next_row, column=6, value=original_comment_viewed)
                worksheet.cell(row=next_row, column=7, value=timestamp)
                
                workbook.save(self.file_path)
                print(f"觀察者動作已記錄: {user_id} - {post_id} - {current_time}")
                return True
            except Exception as e:
                print(f"寫入觀察者動作時出錯: {e}")
                return False
    
    def add_activity(self, user_id, activity_type, page, details=''):
        """添加用戶活動記錄"""
        return self.add_user_entry(user_id, activity_type, page, details)
    
    def get_user_activities(self, user_id):
        """獲取特定用戶的所有活動記錄"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            
            # 檢查是否存在 User_Activity 工作表
            if 'User_Activity' not in workbook.sheetnames:
                print("User_Activity 工作表不存在，返回空活動列表")
                return []
            
            worksheet = workbook['User_Activity']
            
            activities = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row[0] == user_id:  # User ID column
                    activities.append({
                        'user_id': row[0],
                        'created_time': row[1],
                        'activity_type': row[2],
                        'page': row[3],
                        'details': row[4],
                        'timestamp': row[5]
                    })
            return activities
        except Exception as e:
            print(f"讀取用戶活動時出錯: {e}")
            return []
        except Exception as e:
            print(f"讀取Excel時出錯: {e}")
            return []
    
    def get_all_data(self):
        """獲取所有數據"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook.active
            
            data = []
            headers = [cell.value for cell in worksheet[1]]
            
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # 如果行不為空
                    data.append(dict(zip(headers, row)))
            return data
        except Exception as e:
            print(f"讀取所有數據時出錯: {e}")
            return []

# 創建全局實例
excel_manager = ExcelManager()
